"""Command-line entrypoint for the URice Tools Client Python sidecar."""

from __future__ import annotations

import argparse
import contextlib
import io
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import fitz
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from urice_engine import __version__
from urice_engine.core.extraction import POExtractor, configure_tesseract

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

FIELD_HEADERS = {
    "tanggal_diproses": "Tanggal Diproses",
    "company_name": "Company Name",
    "nomor_op": "Nomor OP",
    "nama_vendor": "Nama Vendor",
    "nomor_vendor": "Nomor Vendor",
    "nomor_item": "Nomor Item",
    "nama_barang": "Nama Barang",
    "quantity": "Banyaknya / QTY",
    "satuan": "Satuan",
    "harga_satuan": "Harga Satuan",
    "total": "Total",
    "processed_pdf": "Processed PDF Link",
    "keterangan": "Keterangan",
    "status": "Status",
}

DEFAULT_FIELD_KEYS = [
    "tanggal_diproses",
    "company_name",
    "nomor_op",
    "nama_vendor",
    "nomor_vendor",
    "processed_pdf",
    "keterangan",
    "status",
]

EXCEL_HEADERS = [FIELD_HEADERS[key] for key in DEFAULT_FIELD_KEYS]

HEADER_TO_FIELD = {header: key for key, header in FIELD_HEADERS.items()}

TEXT_COLUMNS = {
    "Nomor OP",
    "Nomor Vendor",
    "Nomor Item",
    "Nomor PO",
    "Nomor PR",
    "Nomor GR",
    "Vendor Code",
    "Material Code",
}

MERGER_STANDARD_COLUMNS = [
    "Source",
    "Nomor PO",
    "Nomor PR",
    "Nomor GR",
    "Vendor Code",
    "Vendor Name",
    "Material Code",
    "Nama Barang",
    "Qty",
    "Satuan",
    "Harga Satuan",
    "Total",
    "Tanggal",
    "Plant",
    "Status",
    "Keterangan",
]


def _excel_text(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text if text.startswith("'") else f"'{text}"


def health() -> dict:
    return {
        "ok": True,
        "engine": "urice-python-sidecar",
        "version": __version__,
        "tesseract_configured": configure_tesseract(),
    }


def _extract_page_payload(path: Path) -> tuple[list[dict], str]:
    doc = fitz.open(str(path))
    try:
        if len(doc) == 0:
            return [], ""
        page = doc[0]
        extractor = POExtractor(log_callback=lambda _message, _level="INFO": None)
        with contextlib.redirect_stdout(io.StringIO()):
            rows = extractor.extract_legacy(page, path.name)
            text = page.get_text("text") or ""
            if len(text.strip()) < 30:
                text = extractor._get_ocr_text(page)
        return rows, _extract_company_name(text)
    finally:
        doc.close()


def _extract_rows(path: Path) -> list[dict]:
    rows, _company = _extract_page_payload(path)
    return rows


def _extract_company_name(text: str) -> str:
    lines = [" ".join(line.strip().split()) for line in text.splitlines() if line.strip()]
    joined = " ".join(lines)
    known_patterns = [
        r"(PT\.?\s+CHAROEN\s+POKPHAND\s+INDONESIA(?:\s+Tbk\.?)?)",
        r"(CHAROEN\s+POKPHAND\s+INDONESIA(?:\s+Tbk\.?)?)",
        r"(PT\.?\s+SINAR\s+TERNAK\s+SEJAHTERA(?:\s*-\s*LAMPUNG)?)",
        r"(SINAR\s+TERNAK\s+SEJAHTERA(?:\s*-\s*LAMPUNG)?)",
        r"(CPI\s+[^\n,]{3,60})",
    ]
    for pattern in known_patterns:
        match = re.search(pattern, joined, re.IGNORECASE)
        if match:
            return " ".join(match.group(1).strip(" -:,").split())[:80]

    for line in lines[:30]:
        upper = line.upper()
        if any(token in upper for token in ["CHAROEN", "SINAR TERNAK", "POKPHAND", "CPI"]):
            if not any(skip in upper for skip in ["VENDOR", "PHONE", "FAX", "EMAIL", "BANK"]):
                return line[:80]
    return ""


def _clean_filename_part(value: str, fallback: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9\s\.\-]", "", value or fallback).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned[:60].strip("_") or fallback


def _save_first_page(source_pdf: Path, output_folder: Path, vendor_name: str, no_op: str) -> tuple[str, str]:
    output_folder.mkdir(parents=True, exist_ok=True)
    filename = f"{_clean_filename_part(vendor_name, 'VENDOR')}-{_clean_filename_part(no_op, datetime.now().strftime('%Y%m%d%H%M%S'))}.pdf"
    output_path = output_folder / filename

    doc = fitz.open(str(source_pdf))
    try:
        if len(doc) == 0:
            raise ValueError("PDF has no pages")
        new_doc = fitz.open()
        try:
            new_doc.insert_pdf(doc, from_page=0, to_page=0)
            new_doc.save(str(output_path), garbage=4, deflate=True)
        finally:
            new_doc.close()
    finally:
        doc.close()

    return str(output_path), filename


def extract_pdf(pdf_path: str) -> dict:
    path = Path(pdf_path).resolve()
    if not path.exists():
        raise FileNotFoundError(str(path))

    rows, company_name = _extract_page_payload(path)
    return {"ok": True, "source": str(path), "company_name": company_name, "rows": rows}


def process_pdf(pdf_path: str, output_folder: str, source_archive_folder: str = "") -> dict:
    path = Path(pdf_path).resolve()
    if not path.exists():
        raise FileNotFoundError(str(path))

    rows, company_name = _extract_page_payload(path)
    if not rows:
        return {"ok": False, "error": "No extraction rows were produced.", "source": str(path), "rows": []}

    first = rows[0]
    processed_path, processed_filename = _save_first_page(
        path,
        Path(output_folder).expanduser().resolve(),
        first.get("vendor_name", "VENDOR"),
        first.get("no_op", ""),
    )

    return {
        "ok": True,
        "source": str(path),
        "processed_pdf_path": processed_path,
        "processed_pdf_filename": processed_filename,
        "archived_source_path": "",
        "company_name": company_name,
        "rows": rows,
    }


def _normalized_field_keys(field_keys: object | None) -> list[str]:
    if not isinstance(field_keys, list):
        return DEFAULT_FIELD_KEYS.copy()
    normalized = [str(key) for key in field_keys if str(key) in FIELD_HEADERS]
    return normalized or DEFAULT_FIELD_KEYS.copy()


def _headers_for_fields(field_keys: object | None) -> list[str]:
    return [FIELD_HEADERS[key] for key in _normalized_field_keys(field_keys)]


def _apply_header_style(ws, headers: list[str]) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="630D16", end_color="630D16", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        if header in TEXT_COLUMNS:
            ws.column_dimensions[get_column_letter(col)].number_format = "@"


def _apply_column_widths(ws, headers: list[str]) -> None:
    widths = {
        "Tanggal Diproses": 20,
        "Company Name": 28,
        "Nomor OP": 18,
        "Nama Vendor": 34,
        "Nomor Vendor": 18,
        "Nomor Item": 16,
        "Nama Barang": 42,
        "Banyaknya / QTY": 18,
        "Satuan": 14,
        "Harga Satuan": 18,
        "Total": 18,
        "Processed PDF Link": 48,
        "Keterangan": 40,
        "Status": 18,
    }
    for index, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(index)].width = widths.get(header, 22)


def ensure_excel_file(excel_path: str, field_keys: object | None = None) -> Path:
    path = Path(excel_path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = _headers_for_fields(field_keys)
        existing_headers = [str(cell.value or "").strip() for cell in ws[1] if str(cell.value or "").strip()]
        if not existing_headers:
            _apply_header_style(ws, headers)
            _apply_column_widths(ws, headers)
        else:
            next_col = len(existing_headers) + 1
            for header in headers:
                if header not in existing_headers:
                    ws.cell(row=1, column=next_col, value=header)
                    existing_headers.append(header)
                    next_col += 1
            _apply_header_style(ws, existing_headers)
            _apply_column_widths(ws, existing_headers)
        ws.freeze_panes = "A2"
        wb.save(path)
        wb.close()
        return path

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data PO"
    headers = _headers_for_fields(field_keys)
    _apply_header_style(ws, headers)
    _apply_column_widths(ws, headers)
    ws.freeze_panes = "A2"
    wb.save(path)
    wb.close()
    return path


def append_record(excel_path: str, record_json: str) -> dict:
    record = json.loads(record_json)
    selected_fields = _normalized_field_keys(record.get("_selected_fields"))
    path = ensure_excel_file(excel_path, selected_fields)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1] if str(cell.value or "").strip()]
    if not headers:
        headers = _headers_for_fields(selected_fields)
        _apply_header_style(ws, headers)
    values = []
    for header in headers:
        field_key = HEADER_TO_FIELD.get(header, "")
        value = record.get(field_key, "")
        if header in TEXT_COLUMNS:
            value = _excel_text(value)
        values.append(value)

    ws.append(values)
    last_row = ws.max_row
    processed_pdf_col = headers.index("Processed PDF Link") + 1 if "Processed PDF Link" in headers else None
    if processed_pdf_col:
        pdf_value = str(record.get("processed_pdf", "") or "").strip()
        if pdf_value:
            cell = ws.cell(row=last_row, column=processed_pdf_col)
            cell.hyperlink = pdf_value
            cell.style = "Hyperlink"

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=last_row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if headers[col - 1] in TEXT_COLUMNS:
            cell.number_format = "@"

    wb.save(path)
    wb.close()
    return {"ok": True, "excel_path": str(path), "row": last_row}


def append_record_file(excel_path: str, record_file: str) -> dict:
    payload = Path(record_file).read_text(encoding="utf-8-sig")
    return append_record(excel_path, payload)


def _sheet_headers(ws) -> list[str]:
    if ws.max_row < 1:
        return []
    headers = []
    for cell in ws[1]:
        text = str(cell.value or "").strip()
        if text:
            headers.append(text)
    return headers


def inspect_excel_headers(excel_path: str) -> dict:
    path = Path(excel_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(str(path))

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = _sheet_headers(ws)
        return {
            "ok": True,
            "path": str(path),
            "sheet": ws.title,
            "headers": headers,
            "row_count": max(ws.max_row - 1, 0),
        }
    finally:
        wb.close()


def _merger_headers_for_sources(sources: list[dict]) -> list[str]:
    selected_targets = set()
    for source in sources:
        mapping = source.get("mapping") or {}
        if not isinstance(mapping, dict):
            continue
        for target in mapping.values():
            target_text = str(target or "").strip()
            if target_text and target_text in MERGER_STANDARD_COLUMNS:
                selected_targets.add(target_text)
    return [header for header in MERGER_STANDARD_COLUMNS if header == "Source" or header in selected_targets]


def _normalize_excel_value(header: str, value: object) -> object:
    if value is None:
        return ""
    if header in TEXT_COLUMNS:
        return _excel_text(value)
    return value


def _style_merger_sheet(ws, headers: list[str]) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="630D16", end_color="630D16", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(index)].width = 22 if header != "Nama Barang" else 42
        if header in TEXT_COLUMNS:
            ws.column_dimensions[get_column_letter(index)].number_format = "@"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if headers[cell.column - 1] in TEXT_COLUMNS:
                cell.number_format = "@"


def merge_excel_files(output_path: str, sources: list[dict]) -> dict:
    if not sources:
        raise ValueError("At least one Excel source is required.")

    output = Path(output_path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    headers = _merger_headers_for_sources(sources)
    if headers == ["Source"]:
        headers = MERGER_STANDARD_COLUMNS.copy()

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Merged Data"
    out_ws.append(headers)

    row_count = 0
    source_summaries = []
    for source in sources:
        source_path = Path(str(source.get("path", ""))).expanduser().resolve()
        source_name = str(source.get("source_name") or source_path.stem or "Source")
        mapping = source.get("mapping") or {}
        if not source_path.exists():
            source_summaries.append({"source_name": source_name, "path": str(source_path), "ok": False, "error": "File not found"})
            continue
        if not isinstance(mapping, dict):
            source_summaries.append({"source_name": source_name, "path": str(source_path), "ok": False, "error": "Mapping must be an object"})
            continue

        wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
        source_rows = 0
        try:
            ws = wb.active
            source_headers = _sheet_headers(ws)
            header_indexes = {header: index for index, header in enumerate(source_headers)}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(value is None or str(value).strip() == "" for value in row):
                    continue
                output_row = []
                for target_header in headers:
                    if target_header == "Source":
                        output_row.append(source_name)
                        continue
                    source_header = next((src for src, target in mapping.items() if str(target).strip() == target_header), "")
                    value = row[header_indexes[source_header]] if source_header in header_indexes and header_indexes[source_header] < len(row) else ""
                    output_row.append(_normalize_excel_value(target_header, value))
                out_ws.append(output_row)
                row_count += 1
                source_rows += 1
        finally:
            wb.close()
        source_summaries.append({"source_name": source_name, "path": str(source_path), "ok": True, "row_count": source_rows})

    _style_merger_sheet(out_ws, headers)
    out_ws.freeze_panes = "A2"
    out_wb.save(output)
    out_wb.close()
    return {
        "ok": True,
        "output_path": str(output),
        "headers": headers,
        "row_count": row_count,
        "sources": source_summaries,
    }

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="URice Tools Client sidecar")
    parser.add_argument("--health", action="store_true", help="Print backend health as JSON")
    parser.add_argument("--extract", metavar="PDF", help="Extract PO data from a PDF and print JSON")
    parser.add_argument("--process", nargs=3, metavar=("PDF", "OUTPUT_FOLDER", "SOURCE_ARCHIVE"), help="Extract PDF and save the first page. Source archive move is handled after Excel apply.")
    parser.add_argument("--ensure-excel", metavar="XLSX", help="Create target Excel if it does not exist")
    parser.add_argument("--ensure-excel-fields", nargs=2, metavar=("XLSX", "FIELDS_JSON"), help="Create or prepare target Excel with selected fields")
    parser.add_argument("--append-record", nargs=2, metavar=("XLSX", "JSON"), help="Append one corrected PO record to Excel")
    parser.add_argument("--append-record-file", nargs=2, metavar=("XLSX", "JSON_FILE"), help="Append one corrected PO record from a JSON file")
    parser.add_argument("--inspect-excel", metavar="XLSX", help="Read first-row headers from an Excel workbook")
    parser.add_argument("--merge-excel-file", nargs=2, metavar=("OUTPUT_XLSX", "SOURCES_JSON_FILE"), help="Merge Excel sources from a JSON file")
    args = parser.parse_args(argv)

    try:
        if args.health:
            payload = health()
        elif args.extract:
            payload = extract_pdf(args.extract)
        elif args.process:
            payload = process_pdf(args.process[0], args.process[1], args.process[2])
        elif args.ensure_excel:
            path = ensure_excel_file(args.ensure_excel)
            payload = {"ok": True, "excel_path": str(path)}
        elif args.ensure_excel_fields:
            field_keys = json.loads(args.ensure_excel_fields[1])
            path = ensure_excel_file(args.ensure_excel_fields[0], field_keys)
            payload = {"ok": True, "excel_path": str(path)}
        elif args.append_record:
            payload = append_record(args.append_record[0], args.append_record[1])
        elif args.append_record_file:
            payload = append_record_file(args.append_record_file[0], args.append_record_file[1])
        elif args.inspect_excel:
            payload = inspect_excel_headers(args.inspect_excel)
        elif args.merge_excel_file:
            sources = json.loads(Path(args.merge_excel_file[1]).read_text(encoding="utf-8-sig"))
            payload = merge_excel_files(args.merge_excel_file[0], sources)
        else:
            payload = {"ok": False, "error": "No command provided"}
            print(json.dumps(payload), file=sys.stderr)
            return 2

        print(json.dumps(payload, ensure_ascii=False))
        return 0
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
