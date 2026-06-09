import sys
from pathlib import Path

import openpyxl

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from engine import inspect_excel_headers, merge_excel_files


def _write_workbook(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Source"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def test_inspect_excel_headers_reads_first_row(tmp_path: Path):
    source = tmp_path / "sap.xlsx"
    _write_workbook(source, ["Purchasing Document", "Material Description", "Order Quantity"], [[45001, "Feed", 12]])

    payload = inspect_excel_headers(str(source))

    assert payload == {
        "ok": True,
        "path": str(source.resolve()),
        "sheet": "Source",
        "headers": ["Purchasing Document", "Material Description", "Order Quantity"],
        "row_count": 1,
    }


def test_merge_excel_files_normalizes_selected_mapping_and_preserves_text(tmp_path: Path):
    source = tmp_path / "sap.xlsx"
    output = tmp_path / "merged.xlsx"
    _write_workbook(
        source,
        ["Purchasing Document", "Material Description", "Order Quantity", "Vendor"],
        [["0004500123", "Feed premix", 12, "0007000940"]],
    )
    mapping = {
        "Purchasing Document": "Nomor PO",
        "Material Description": "Nama Barang",
        "Order Quantity": "Qty",
        "Vendor": "Vendor Code",
    }

    payload = merge_excel_files(
        output_path=str(output),
        sources=[{"path": str(source), "source_name": "SAP", "mapping": mapping}],
    )

    assert payload["ok"] is True
    assert payload["row_count"] == 1

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    assert [cell.value for cell in ws[1]] == ["Source", "Nomor PO", "Vendor Code", "Nama Barang", "Qty"]
    assert ws.cell(row=2, column=1).value == "SAP"
    assert ws.cell(row=2, column=2).value == "'0004500123"
    assert ws.cell(row=2, column=3).value == "'0007000940"
    assert ws.cell(row=2, column=4).value == "Feed premix"
    assert ws.cell(row=2, column=5).value == 12
    wb.close()
