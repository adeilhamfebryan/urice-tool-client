"""
Excel Service for Full Client.

Handles creating and appending to the target Excel file.
Cleaned up from prototype for better maintainability.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict

from ..models.po_data import POExtractionResult


class ExcelService:
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path).resolve()

    def ensure_file_exists(self):
        """Create the Excel file with proper headers if it doesn't exist."""
        if self.excel_path.exists():
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data PO"

        headers = [
            "Tanggal Proses", "No OP", "Vendor Name", "Vendor Code",
            "Item", "Nama Barang", "Banyaknya", "Satuan",
            "Harga Satuan", "Total", "Grand Total", "Processed PDF"
        ]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='B4B4B4'),
            right=Side(style='thin', color='B4B4B4'),
            top=Side(style='thin', color='B4B4B4'),
            bottom=Side(style='thin', color='B4B4B4')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Reasonable column widths
        widths = [15, 13, 22, 14, 16, 38, 11, 9, 14, 14, 16, 28]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"
        wb.save(self.excel_path)
        wb.close()

    def append_results(self, results: List[POExtractionResult], processed_files: List[Dict]):
        """
        Append multiple extraction results.
        processed_files should contain mapping like {"result": ..., "filename": ..., "path": ...}
        """
        self.ensure_file_exists()

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active

        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        for res in results:
            rows = res.to_excel_rows()
            for row_data in rows:
                # Find the matching processed file info
                fname = ""
                ppath = ""
                for pf in processed_files:
                    if pf.get("result") is res:
                        fname = pf.get("filename", "")
                        ppath = pf.get("path", "")
                        break

                excel_row = [
                    row_data.get("tanggal_proses", ""),
                    row_data.get("no_op", ""),
                    row_data.get("vendor_name", ""),
                    row_data.get("vendor_code", ""),
                    row_data.get("item", ""),
                    row_data.get("nama_barang", ""),
                    row_data.get("quantity", ""),
                    row_data.get("satuan", ""),
                    row_data.get("harga_satuan", ""),
                    row_data.get("total", ""),
                    row_data.get("grand_total", ""),
                    fname
                ]

                ws.append(excel_row)
                last_row = ws.max_row

                # Add hyperlink if we have the processed PDF path
                if ppath and Path(ppath).exists():
                    cell = ws.cell(row=last_row, column=12)
                    abs_path = str(Path(ppath).resolve()).replace("\\", "/")
                    cell.hyperlink = f"file:///{abs_path}"
                    cell.value = fname
                    cell.font = Font(color="0563C1", underline="single")

                for col in range(1, 13):
                    cell = ws.cell(row=last_row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

        wb.save(self.excel_path)
        wb.close()
